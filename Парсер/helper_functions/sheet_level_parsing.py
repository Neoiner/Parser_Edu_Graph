from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import cell as xlcell
from consts.days_of_week import weekdays
from consts.lecture_time_ranges import lecture_time_ranges, first_time_range, second_time_range, third_time_range
from consts.lecture_time_ranges import fourth_time_range, fifth_time_range, sixth_time_range, seventh_time_range
from db_api.model.schedule_group_type_day import ScheduleGroupTypeOneEvent
import urllib.request
import re
import datetime

async def parse_worksheet(worksheet: Worksheet, academic_degree, faculty, year, dryrun=True):
    start_column_index, end_column_index, groupname_row_index = get_column_indexes_and_groupname_row_index(worksheet)
    arr_index = get_start_index_of_course(worksheet)


    curr_course = 1
    end_course = len(arr_index)

    for one_course in arr_index:
        #print(worksheet.cell(row=worksheet.max_row, column=1))
        if curr_course == end_course:
            await parse_for_one_course(worksheet=worksheet, groupname_row_index=groupname_row_index,
                                       start_row_index=one_course,
                                       end_row_index=worksheet.cell(row=worksheet.max_row, column=1),
                                       academic_degree=academic_degree, faculty=faculty, year=year, dryrun=dryrun,
                                       curr_course=curr_course, last_value=1)
        else:
            await parse_for_one_course(worksheet=worksheet, groupname_row_index=groupname_row_index,
                                       start_row_index=one_course,
                                       end_row_index=arr_index[curr_course],
                                       academic_degree=academic_degree, faculty=faculty, year=year,
                                       curr_course=curr_course, dryrun=dryrun)
            curr_course += 1




async def parse_for_one_course(worksheet: Worksheet, groupname_row_index, start_row_index, end_row_index,
                                       academic_degree, faculty, year, curr_course, dryrun=True, last_value=0):
    column_index = 1
    attendee_str_raw = str(worksheet.cell(row=groupname_row_index, column=column_index).value)

    column_index = 1

    course_event = []
    last_term_int_raw = 1

    for index_row in range(start_row_index.row + 1, end_row_index.row + last_value):
        if str(worksheet.cell(row=index_row, column=column_index).value) != "None":
            term_int_raw = worksheet.cell(row=index_row, column=1).value
            last_term_int_raw = term_int_raw
        else:
            term_int_raw = last_term_int_raw
        format_str_raw_tmp = str(worksheet.cell(row=index_row, column=2).value)
        format_str_raw = ""
        if format_str_raw_tmp.find("Теоретическое обучение") != -1:
            format_str_raw = "study"
        elif (format_str_raw_tmp.find("сессия") != -1):
            format_str_raw = "exams"
        elif (format_str_raw_tmp.find("Повторная промежуточная аттестация") != -1):
            format_str_raw = "attestation"
        elif format_str_raw_tmp.find("Каникулы") != -1:
            format_str_raw = "holidays"
        elif format_str_raw_tmp.find("Нерабочие праздничные дни") != -1:
            format_str_raw = "weekends"
        elif format_str_raw_tmp.find("практика") != -1 or format_str_raw_tmp.find("Практика") != -1:
            format_str_raw = "practice"
        elif format_str_raw_tmp.find("Подготовка к сдаче и сдача государственной итоговой аттестации") != -1 or\
                format_str_raw_tmp.find("Период работы ГЭК") != -1:
            format_str_raw = "graduate_work"
        elif format_str_raw_tmp.find("Архитектурные дни") != -1:
            format_str_raw = "archday"
        else:
            format_str_raw = "Unknown"
        summary_str_raw = format_str_raw_tmp

        start_date_raw = str(worksheet.cell(row=index_row, column=3).value)
        end_date_raw = str(worksheet.cell(row=index_row, column=4).value)

        group_fullname_raw = str(worksheet)[str(worksheet).find('"')+1:str(worksheet).rfind('"')]


        if format_str_raw == "weekends" or format_str_raw == "archday":
            date_holidays = (str(worksheet.cell(row=index_row, column=3).value)).split("\n")

            for day in date_holidays:
                if day.find("-", 0, len(day)) != -1:
                    start_date_raw = day.split("-")[0].replace(" ", "").replace("\n", "")
                    end_date_raw = day.split("-")[1].replace(" ", "").replace("\n", "")
                    if start_date_raw == "None":
                        continue
                else:
                    start_date_raw = day.replace(" ", "").replace("\n", "")
                    end_date_raw = day.replace(" ", "").replace("\n", "")
                    if start_date_raw == "None":
                        continue
                schedule_group_event = ScheduleGroupTypeOneEvent(
                    speciality=attendee_str_raw,
                    format=format_str_raw,
                    term=term_int_raw,
                    summary=summary_str_raw.replace("\n", ""),
                    start=datetime.datetime.strptime(start_date_raw.replace(".", "/"), "%d/%m/%Y"),
                    end=datetime.datetime.strptime(end_date_raw.replace(".", "/"), "%d/%m/%Y"),
                    course=curr_course,
                    group_fullname=group_fullname_raw
                )
                if dryrun:
                    print(schedule_group_event)
                else:
                    await schedule_group_event.create()
        else:
            if start_date_raw == "None":
                continue
            schedule_group_event = ScheduleGroupTypeOneEvent(
                speciality=attendee_str_raw,
                format=format_str_raw,
                term=term_int_raw,
                summary=summary_str_raw.replace("\n", ""),
                start=datetime.datetime.strptime(start_date_raw.split(" ")[0].replace("-", "/").replace(".", "/").replace("\n", ""), "%Y/%m/%d"),
                end=datetime.datetime.strptime(end_date_raw.split(" ")[0].replace("-", "/").replace(".", "/").replace("\n", ""), "%Y/%m/%d"),
                course=curr_course,
                group_fullname=group_fullname_raw
            )

            if dryrun:
                print(schedule_group_event)
            else:
                await schedule_group_event.create()



# section get cell value
def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row >= row_start and row <= row_end:
        column = cell.column
        if column >= column_start and column <= column_end:
            return True
    return False


def get_cell_value(sheet: Worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value



# section indexes
def get_column_indexes_and_groupname_row_index(target_worksheet: Worksheet):
    for row in target_worksheet.iter_rows():
        for cell in row:
            if str(cell.value).lower().startswith("информация по периодам обучения обучающихся"):
                groupname_row_index = cell.column + 1
                starting_column_index = cell.column + 3
                ending_column_index = target_worksheet.max_column
                return starting_column_index, ending_column_index, groupname_row_index



def get_start_index_of_course(target_worksheet: Worksheet):

    FIRST_COURSE = "1-й курс"
    SECOND_COURSE = "2-й курс"
    THIRD_COURSE = "3-й курс"
    CHETV_COURSE = "4-й курс"
    PIAT_COURSE = "5-й курс"
    arr_index = []
    for row in target_worksheet.iter_rows():
        for cell in row:
            cell_value = get_cell_value(target_worksheet, cell)
            if (cell_value == FIRST_COURSE) or (cell_value == SECOND_COURSE) or (cell_value == THIRD_COURSE) or (cell_value == CHETV_COURSE) or (cell_value == PIAT_COURSE):
                arr_index.append(cell)
                break

    return arr_index

async def save_graf(saving_directory, url_graf):
    for curr_url in url_graf:
        urllib.request.urlretrieve(curr_url, saving_directory)

